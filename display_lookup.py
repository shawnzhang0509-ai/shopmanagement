"""Display 库数据层：抓取程序写 data/display.xlsx，模板编辑器只读可视化。

架构（和库存 main_gui 项目一样）：
  grab_display.bat        → Main：SQL 抓数据 → data/display.xlsx
  start_furniture_sim.bat → 可视化：furniture_sim 读 Excel 显示 Display 大库
"""
from __future__ import annotations

import json
import os
import re
from dataclasses import dataclass, field
from urllib.parse import quote_plus
from datetime import datetime, timezone
from typing import Any

SCRIPT_DIR = os.path.dirname(os.path.abspath(__file__))
GRABBER_CONFIG = os.path.join(SCRIPT_DIR, "grabber_config.json")
GRABBER_CONFIG_EXAMPLE = os.path.join(SCRIPT_DIR, "grabber_config.example.json")
CACHE_FILE = os.path.join(SCRIPT_DIR, "display_cache.json")
DEFAULT_EXCEL = os.path.join(SCRIPT_DIR, "data", "display.xlsx")
LEGACY_EXCEL = os.path.join(SCRIPT_DIR, "display.xlsx")
DEFAULT_SQL = os.path.join(SCRIPT_DIR, "sql", "display.sql")
DEFAULT_BLACKLIST = os.path.join(SCRIPT_DIR, "data", "display_blacklist.xlsx")
DEFAULT_BLACKLIST_CSV = os.path.join(SCRIPT_DIR, "data", "display_blacklist.csv")
EXAMPLE_BLACKLIST_CSV = os.path.join(SCRIPT_DIR, "data", "display_blacklist.example.csv")

# 门店：按 Stock Details 里的 location 名称匹配
SHOPS: list[dict[str, Any]] = [
    {"id": "all", "label": "全部", "patterns": []},
    {"id": "onehunga", "label": "Onehunga", "patterns": ["onehunga"]},
    {"id": "westgate", "label": "Westgate", "patterns": ["westgate"]},
    {"id": "hamilton", "label": "Hamilton", "patterns": ["hamilton"]},
    {"id": "chch", "label": "Christchurch", "patterns": ["chch", "christchurch", "colombo"]},
    {"id": "carbine", "label": "Carbine Rd", "patterns": ["carbine"]},
    {"id": "other", "label": "其他", "patterns": []},
]

# Excel / SQL 列名别名（不区分大小写，空格会折叠）
_COL_ALIASES: dict[str, tuple[str, ...]] = {
    "warehouse_name": ("warehouse_name", "warehousename", "warehouse name", "warehouse"),
    "product_code": ("product_code", "productcode", "code", "sku", "product code", "item code"),
    "product_name": ("product_name", "productname", "product name", "title", "description"),
    "product_family": (
        "product_family",
        "productfamily",
        "family",
        "product family",
        "family name",
        "familyname",
        "range",
        "collection",
        "产品系列",
        "系列",
    ),
    "sub_product_family": (
        "sub_product_family",
        "subproductfamily",
        "sub product family",
        "subfamily",
        "sub family",
        "product line",
        "line",
    ),
    "display_qty": ("display_qty", "displayqty", "display qty", "quantity", "qty"),
    "image_url": (
        "image_url",
        "imageurl",
        "image url",
        "imagepath",
        "image path",
        "image",
        "picture",
        "photo",
        "thumbnail",
    ),
    "stock_details": ("stock_details", "stockdetails", "stock details", "stock", "display stock", "inventory"),
}

_display_cache: list["DisplayItem"] | None = None
_display_items_all: list["DisplayItem"] | None = None
_last_load_error: str | None = None
_last_load_source: str | None = None
_last_family_column: str | None = None
_last_sql_file: str | None = None


@dataclass
class DisplaySlot:
    shop_id: str
    shop_label: str
    location: str
    qty: int


@dataclass
class DisplayItem:
    product_code: str
    product_name: str
    product_family: str
    sub_product_family: str = ""
    image_url: str = ""
    stock_details: str = ""
    displays: list[DisplaySlot] = field(default_factory=list)

    @property
    def key(self) -> str:
        return self.product_code or self.product_name

    def display_qty_for_shop(self, shop_id: str) -> int:
        if shop_id == "all":
            return sum(s.qty for s in self.displays)
        return sum(s.qty for s in self.displays if s.shop_id == shop_id)

    def shops_with_display(self) -> set[str]:
        return {s.shop_id for s in self.displays if s.qty > 0}


def _normalize_key(value: str) -> str:
    return re.sub(r"\s+", " ", str(value or "").strip().lower())


def _normalize_header(value) -> str:
    return re.sub(r"\s+", " ", str(value or "").strip().lower())


def _cell_value(val) -> str:
    """Excel/SQL 单元格 → 字符串；None / NaN / 'nan' 视为空。"""
    if val is None:
        return ""
    if isinstance(val, float):
        import math

        if math.isnan(val):
            return ""
    s = str(val).strip()
    if s.lower() in ("nan", "none", "null", "#n/a", "n/a"):
        return ""
    return s


def _canonicalize_row(raw: dict) -> dict:
    """把 Excel / SQL 任意列名映射到标准字段名。"""
    normalized = {_normalize_header(k): v for k, v in raw.items()}
    out: dict = {}
    for field, aliases in _COL_ALIASES.items():
        for alias in aliases:
            if alias in normalized:
                out[field] = _cell_value(normalized[alias])
                break
    return out


def _header_is_sub_family(header: str) -> bool:
    h = _normalize_header(header)
    compact = h.replace(" ", "")
    return h.startswith("sub") or " sub" in f" {h} " or compact.startswith("subproduct")


def _header_is_product_family(header: str) -> bool:
    h = _normalize_header(header)
    if _header_is_sub_family(h):
        return False
    if h in _COL_ALIASES["product_family"]:
        return True
    compact = h.replace(" ", "")
    if compact in ("productfamily", "familyname", "family"):
        return True
    if "系列" in h:
        return True
    if h.endswith("family") and "name" not in h.replace("family", ""):
        return True
    return False


def _infer_family_column_indices(headers: list[str]) -> tuple[int | None, int | None]:
    family_idx = sub_idx = None
    for i, h in enumerate(headers):
        if sub_idx is None and _header_is_sub_family(h):
            sub_idx = i
    for i, h in enumerate(headers):
        if family_idx is None and _header_is_product_family(h):
            family_idx = i
    return family_idx, sub_idx


def _resolve_family_name(family: str, name: str, code: str) -> str:
    """只用 Excel/SQL 的 ProductFamily；缺失时标为未分类，不用 SKU/产品名猜测。"""
    if family:
        return family
    return "未分类"


def _resolve_sub_family_name(sub_family: str, name: str, code: str) -> str:
    if sub_family:
        return sub_family
    return name or code or "未分类"


def _family_data_score(items: list[DisplayItem]) -> int:
    """优先选用真正带有 ProductFamily 数据的 Excel。"""
    score = 0
    for it in items:
        fam = it.product_family or ""
        if not fam or fam == "未分类":
            continue
        if fam == it.product_code:
            continue
        first = (it.product_name or "").split(" ")[0]
        if fam == first and re.match(r"^\d{3}-\d{3}$", fam):
            continue
        score += 1
    return score


def _detect_format(fields: dict) -> str | None:
    keys = set(fields)
    has_product = "product_code" in keys or "product_name" in keys
    if "warehouse_name" in keys and "display_qty" in keys and has_product:
        return "warehouse"
    if "stock_details" in keys and has_product:
        return "stock"
    return None


def _resolve_columns(headers: list[str]) -> tuple[str, dict[str, int]] | None:
    sample = {}
    for field, aliases in _COL_ALIASES.items():
        for i, h in enumerate(headers):
            if h in aliases:
                sample[field] = True
                break
    fmt = _detect_format(sample)
    if not fmt:
        return None
    mapping: dict[str, int] = {}
    for field, aliases in _COL_ALIASES.items():
        for i, h in enumerate(headers):
            if h in aliases:
                mapping[field] = i
                break
    # 宽松匹配 family 列
    for i, h in enumerate(headers):
        if "sub_product_family" not in mapping and _header_is_sub_family(h):
            mapping["sub_product_family"] = i
        elif "product_family" not in mapping and _header_is_product_family(h):
            mapping["product_family"] = i
    return fmt, mapping


def is_display_location(location: str) -> bool:
    low = location.lower()
    if "no longer available" in low:
        return False
    return "display" in low


def shop_id_for_location(location: str) -> str:
    low = location.lower()
    for shop in SHOPS:
        if shop["id"] in ("all", "other"):
            continue
        if any(p in low for p in shop["patterns"]):
            return shop["id"]
    return "other"


def shop_label(shop_id: str) -> str:
    for shop in SHOPS:
        if shop["id"] == shop_id:
            return shop["label"]
    return shop_id


def parse_stock_details(text: str) -> list[DisplaySlot]:
    slots: list[DisplaySlot] = []
    if not text:
        return slots
    for part in str(text).split(";"):
        part = part.strip()
        if not part or ":" not in part:
            continue
        loc, qty_s = part.rsplit(":", 1)
        loc = loc.strip()
        if not is_display_location(loc):
            continue
        try:
            qty = int(float(qty_s.strip()))
        except (TypeError, ValueError):
            qty = 0
        if qty <= 0:
            continue
        sid = shop_id_for_location(loc)
        slots.append(DisplaySlot(sid, shop_label(sid), loc, qty))
    return slots


def _cell_str(row: tuple, idx: int | None) -> str:
    if idx is None or idx >= len(row):
        return ""
    return _cell_value(row[idx])


def _row_to_item(row: dict) -> DisplayItem | None:
    code = _cell_value(row.get("product_code") or row.get("sku") or row.get("code"))
    name = _cell_value(row.get("product_name") or row.get("name") or row.get("title"))
    family = _cell_value(row.get("product_family") or row.get("family"))
    sub_family = _cell_value(row.get("sub_product_family") or row.get("subfamily"))
    image_url = _cell_value(row.get("image_url") or row.get("imageurl") or row.get("imagepath"))
    stock = _cell_value(row.get("stock_details") or row.get("stock") or row.get("Stock Details"))
    if not name and not code:
        return None
    family = _resolve_family_name(family, name, code)
    sub_family = _resolve_sub_family_name(sub_family, name, code)
    displays = row.get("displays")
    if isinstance(displays, list) and displays:
        slots = [
            DisplaySlot(
                str(d.get("shop_id", "other")),
                str(d.get("shop_label", shop_label(str(d.get("shop_id", "other"))))),
                str(d.get("location", "")),
                int(d.get("qty", 0) or 0),
            )
            for d in displays
            if int(d.get("qty", 0) or 0) > 0
        ]
    else:
        slots = parse_stock_details(stock)
    if not slots:
        return None
    return DisplayItem(code or name, name or code, family, sub_family, image_url, stock, slots)


def _aggregate_warehouse_rows(rows: list[dict]) -> list[DisplayItem]:
    """多行（每行一个仓库+产品）合并为 DisplayItem。"""
    groups: dict[str, dict] = {}
    for raw in rows:
        row = raw if "warehouse_name" in raw else _canonicalize_row(raw)
        code = _cell_value(row.get("product_code"))
        name = _cell_value(row.get("product_name"))
        warehouse = _cell_value(row.get("warehouse_name"))
        try:
            qty = int(float(row.get("display_qty") or 0))
        except (TypeError, ValueError):
            qty = 0
        if not warehouse or qty <= 0 or (not code and not name):
            continue
        key = _normalize_key(code or name)
        family = _cell_value(row.get("product_family"))
        sub_family = _cell_value(row.get("sub_product_family"))
        image_url = _cell_value(row.get("image_url"))
        if key not in groups:
            groups[key] = {
                "product_code": code or name,
                "product_name": name or code,
                "product_family": _resolve_family_name(family, name, code),
                "sub_product_family": _resolve_sub_family_name(sub_family, name, code),
                "image_url": image_url,
                "slots": {},
            }
        else:
            if family:
                groups[key]["product_family"] = _resolve_family_name(
                    family, groups[key]["product_name"], groups[key]["product_code"]
                )
            if sub_family:
                groups[key]["sub_product_family"] = sub_family
            if image_url and not groups[key].get("image_url"):
                groups[key]["image_url"] = image_url
        sid = shop_id_for_location(warehouse)
        slot_key = (sid, warehouse)
        groups[key]["slots"][slot_key] = groups[key]["slots"].get(slot_key, 0) + qty

    items: list[DisplayItem] = []
    for g in groups.values():
        displays = [
            {"shop_id": sid, "shop_label": shop_label(sid), "location": loc, "qty": q}
            for (sid, loc), q in g["slots"].items()
        ]
        stock = ";".join(f"{d['location']}:{d['qty']}" for d in displays)
        item = _row_to_item({**g, "displays": displays, "stock_details": stock})
        if item:
            items.append(item)
    return items


def _rows_to_items(rows: list[dict], fmt: str | None = None, *, apply_blacklist: bool = True) -> list[DisplayItem]:
    if not rows:
        return []
    canonical = [_canonicalize_row(r) for r in rows]
    if fmt is None:
        fmt = _detect_format(canonical[0])
    if fmt == "warehouse":
        items = _aggregate_warehouse_rows(canonical)
    else:
        items = []
        for row in canonical:
            item = _row_to_item(row)
            if item:
                items.append(item)
    if apply_blacklist:
        return filter_blacklisted(items)
    return items


_BLACKLIST_ALIASES = {
    "sku": ("sku", "product_code", "productcode", "code", "product code", "item code"),
    "product_name": ("product_name", "productname", "name", "product name", "title"),
}


def resolve_blacklist_paths(cfg: dict | None = None) -> list[str]:
    cfg = cfg or load_grabber_config()
    paths: list[str] = []
    if cfg.get("blacklist_file"):
        paths.append(_resolve_path(cfg["blacklist_file"]))
    paths.extend([
        DEFAULT_BLACKLIST,
        DEFAULT_BLACKLIST_CSV,
        EXAMPLE_BLACKLIST_CSV,
        os.path.join(SCRIPT_DIR, "display_blacklist.xlsx"),
    ])
    seen: set[str] = set()
    out: list[str] = []
    for p in paths:
        if p not in seen:
            seen.add(p)
            out.append(p)
    return out


def blacklist_files_revision() -> str:
    """黑名单文件修改时间（画廊布局缓存用，改 CSV 后自动刷新）。"""
    parts: list[str] = []
    for path in resolve_blacklist_paths():
        if os.path.isfile(path):
            parts.append(f"{os.path.basename(path)}:{int(os.path.getmtime(path))}")
    return "|".join(parts)


def _blacklist_key(value) -> str:
    """黑名单 SKU 归一化（去空格、Excel 数字格式）。"""
    s = _cell_value(value)
    if not s:
        return ""
    if re.fullmatch(r"\d+\.0+", s):
        s = s.split(".", 1)[0]
    return _normalize_key(s)


def _blacklist_add_sku(blocked: set[str], value) -> None:
    key = _blacklist_key(value)
    if key:
        blocked.add(key)
        compact = re.sub(r"[^a-z0-9]", "", key)
        if compact:
            blocked.add(compact)


def _blacklist_read_csv_rows(path: str) -> list[list[str]]:
    import csv

    last_err: Exception | None = None
    for encoding in ("utf-8-sig", "utf-8", "gbk", "cp936", "latin-1"):
        try:
            with open(path, "r", encoding=encoding, newline="") as f:
                return list(csv.reader(f))
        except UnicodeDecodeError as exc:
            last_err = exc
            continue
    if last_err is not None:
        raise last_err
    return []


def load_blacklist(cfg: dict | None = None) -> set[str]:
    """读取黑名单 SKU（支持仅一列 SKU，也兼容产品名列）。"""
    blocked: set[str] = set()
    for path in resolve_blacklist_paths(cfg):
        if not os.path.isfile(path):
            continue
        if path.lower().endswith(".csv"):
            rows = _blacklist_read_csv_rows(path)
            if not rows:
                continue
            headers = [_normalize_header(h) for h in rows[0]]
            sku_col = name_col = None
            for i, h in enumerate(headers):
                if h in _BLACKLIST_ALIASES["sku"]:
                    sku_col = i
                if h in _BLACKLIST_ALIASES["product_name"]:
                    name_col = i
            data_start = 1
            # 无表头：整文件每行第一列当作 SKU
            if sku_col is None and name_col is None:
                if len(headers) == 1:
                    sku_col = 0
                    data_start = 0
                elif all(_normalize_key(h) for h in rows[0]):
                    sku_col = 0
                    data_start = 0
            elif sku_col is None and len(headers) == 1:
                sku_col = 0
            for line in rows[data_start:]:
                if not line:
                    continue
                if sku_col is not None and sku_col < len(line):
                    _blacklist_add_sku(blocked, line[sku_col])
                if name_col is not None and name_col < len(line):
                    _blacklist_add_sku(blocked, line[name_col])
            continue
        try:
            import pandas as pd

            df = pd.read_excel(path)
            df.columns = [_normalize_header(c) for c in df.columns]
            for field, aliases in _BLACKLIST_ALIASES.items():
                for alias in aliases:
                    if alias in df.columns:
                        for val in df[alias].dropna():
                            key = _normalize_key(str(val))
                            if key:
                                blocked.add(key)
                        break
            continue
        except ImportError:
            pass
        from openpyxl import load_workbook

        wb = load_workbook(path, read_only=True, data_only=True)
        ws = wb.active
        raw = list(ws.iter_rows(values_only=True))
        wb.close()
        if not raw:
            continue
        headers = [_normalize_header(h) for h in raw[0]]
        sku_col = name_col = None
        for i, h in enumerate(headers):
            if h in _BLACKLIST_ALIASES["sku"]:
                sku_col = i
            if h in _BLACKLIST_ALIASES["product_name"]:
                name_col = i
        data_start = 1
        if sku_col is None and name_col is None and len(headers) == 1:
            sku_col = 0
            data_start = 0
        elif sku_col is None and len(headers) == 1:
            sku_col = 0
        for line in raw[data_start:]:
            if sku_col is not None and sku_col < len(line) and line[sku_col]:
                _blacklist_add_sku(blocked, line[sku_col])
            if name_col is not None and name_col < len(line) and line[name_col]:
                _blacklist_add_sku(blocked, line[name_col])
    return blocked


def is_blacklisted(item: DisplayItem, blocked: set[str]) -> bool:
    if not blocked:
        return False
    keys = {_blacklist_key(item.product_code), _blacklist_key(item.product_name)}
    keys.discard("")
    for key in keys:
        if key in blocked:
            return True
        compact = re.sub(r"[^a-z0-9]", "", key)
        if compact and compact in blocked:
            return True
    return False


def blacklist_status(cfg: dict | None = None) -> tuple[int, str, list[str]]:
    """返回 (SKU 数量, 主文件名, 实际读到的文件列表)。"""
    blocked = load_blacklist(cfg)
    used = [p for p in resolve_blacklist_paths(cfg) if os.path.isfile(p)]
    names = [os.path.basename(p) for p in used]
    primary = names[0] if names else "未找到黑名单文件"
    return len(blocked), primary, names


def filter_blacklisted(items: list[DisplayItem], cfg: dict | None = None) -> list[DisplayItem]:
    blocked = load_blacklist(cfg)
    if not blocked:
        return items
    return [it for it in items if not is_blacklisted(it, blocked)]


def _apply_family_fields(
    row: dict,
    values: tuple,
    headers: list[str],
    family_idx: int | None,
    sub_idx: int | None,
) -> None:
    if family_idx is not None and family_idx < len(values):
        row["product_family"] = _cell_value(values[family_idx])
    if sub_idx is not None and sub_idx < len(values):
        row["sub_product_family"] = _cell_value(values[sub_idx])


def _load_excel_rows(path: str) -> tuple[str, list[dict]]:
    global _last_family_column
    if not os.path.isfile(path):
        return "warehouse", []

    try:
        import pandas as pd

        df = pd.read_excel(path)
        df.columns = [_normalize_header(c) for c in df.columns]
        headers = list(df.columns)
        resolved = _resolve_columns(headers)
        if not resolved:
            raise ValueError(
                f"{os.path.basename(path)} 列不匹配。需要 WarehouseName+Sku+ProductName+DisplayQty（可选 ProductFamily/SubProductFamily），"
                "或 stock_details + product_name"
            )
        fmt, colmap = resolved
        family_idx, sub_idx = _infer_family_column_indices(headers)
        if family_idx is not None:
            _last_family_column = headers[family_idx]
        elif "product_family" in colmap:
            _last_family_column = headers[colmap["product_family"]]
        rows: list[dict] = []
        for _, series in df.iterrows():
            row = {
                field: _cell_value(series.iloc[idx] if idx < len(series) else None)
                for field, idx in colmap.items()
            }
            _apply_family_fields(row, tuple(series.tolist()), headers, family_idx, sub_idx)
            rows.append(row)
        return fmt, rows
    except ImportError:
        pass

    from openpyxl import load_workbook

    wb = load_workbook(path, read_only=True, data_only=True)
    ws = wb.active
    raw = list(ws.iter_rows(values_only=True))
    wb.close()
    if not raw:
        return "warehouse", []
    headers = [_normalize_header(h) for h in raw[0]]
    resolved = _resolve_columns(headers)
    if not resolved:
        raise ValueError(
            f"{os.path.basename(path)} 列不匹配。需要 WarehouseName+Sku+ProductName+DisplayQty，"
            "或 stock_details + product_name"
        )
    fmt, colmap = resolved
    family_idx, sub_idx = _infer_family_column_indices(headers)
    if family_idx is not None:
        _last_family_column = headers[family_idx]
    elif "product_family" in colmap:
        _last_family_column = headers[colmap["product_family"]]
    rows: list[dict] = []
    for line in raw[1:]:
        if not line:
            continue
        row = {field: _cell_str(line, idx) for field, idx in colmap.items()}
        _apply_family_fields(row, line, headers, family_idx, sub_idx)
        rows.append(row)
    return fmt, rows


def load_from_excel(path: str | None = None) -> list[DisplayItem]:
    global _last_load_error, _last_load_source, _last_family_column
    candidates = [path] if path else resolve_display_excel_paths()
    last_exc = None
    best_items: list[DisplayItem] = []
    best_score = -1
    best_source: str | None = None
    for candidate in candidates:
        if not candidate or not os.path.isfile(candidate):
            continue
        try:
            fmt, rows = _load_excel_rows(candidate)
            items_all = _rows_to_items(rows, fmt, apply_blacklist=False)
            global _display_items_all
            _display_items_all = items_all
            items = filter_blacklisted(items_all)
            if not items:
                _last_load_error = f"{os.path.basename(candidate)} 中没有 Display 数据"
                continue
            score = _family_data_score(items)
            if score > best_score:
                best_score = score
                best_items = items
                best_source = candidate
        except Exception as exc:
            last_exc = exc
            _last_load_error = f"读取 {os.path.basename(candidate)} 失败: {exc}"
    if best_items and best_source:
        _last_load_source = os.path.basename(best_source)
        if best_score <= 0:
            col = _last_family_column or "ProductFamily"
            _last_load_error = (
                f"未读到有效的 {col} 数据，界面将显示「未分类」。"
                "请确认 Excel 有 ProductFamily 列且已填写，或重新 grab_display.bat 抓取。"
            )
        else:
            _last_load_error = None
        return best_items
    if not candidates or not any(os.path.isfile(p) for p in candidates if p):
        _last_load_error = "未找到 display.xlsx，请先运行 grab_display.bat 抓取数据"
    elif last_exc:
        _last_load_error = str(last_exc)
    _last_load_source = None
    return []


def load_grabber_config() -> dict:
    for path in (GRABBER_CONFIG, GRABBER_CONFIG_EXAMPLE):
        if os.path.isfile(path):
            with open(path, "r", encoding="utf-8") as f:
                return json.load(f)
    return {}


def save_grabber_config(cfg: dict) -> None:
    path = GRABBER_CONFIG
    existing: dict = {}
    if os.path.isfile(path):
        try:
            with open(path, "r", encoding="utf-8") as f:
                existing = json.load(f)
        except (OSError, json.JSONDecodeError):
            existing = {}
    merged = {**existing, **cfg}
    with open(path, "w", encoding="utf-8") as f:
        json.dump(merged, f, ensure_ascii=False, indent=2)


def _resolve_path(path: str) -> str:
    if not path:
        return SCRIPT_DIR
    return path if os.path.isabs(path) else os.path.join(SCRIPT_DIR, path)


def build_runtime_config(cfg: dict | None = None) -> dict:
    """合并配置并解析 sql / 输出路径。"""
    base = dict(cfg or load_grabber_config())
    sql_folder = base.get("sql_folder") or "sql"
    sql_folder_abs = _resolve_path(sql_folder)
    sql_file = base.get("sql_file")
    if not sql_file:
        sql_file = os.path.join(sql_folder_abs, "display.sql")
    else:
        sql_file = _resolve_path(sql_file)
    base["sql_file"] = sql_file

    output_folder = base.get("output_folder") or "data"
    output_folder_abs = _resolve_path(output_folder)
    if not base.get("output_excel"):
        base["output_excel"] = os.path.join(output_folder_abs, "display.xlsx")
    else:
        base["output_excel"] = _resolve_path(base["output_excel"])
    if not base.get("output_json"):
        base["output_json"] = os.path.join(output_folder_abs, "display_cache.json")
    else:
        base["output_json"] = _resolve_path(base["output_json"])
    return base


def resolve_display_excel_paths() -> list[str]:
    cfg = load_grabber_config()
    paths: list[str] = []
    if cfg.get("output_excel"):
        paths.append(os.path.join(SCRIPT_DIR, cfg["output_excel"]))
    paths.extend([DEFAULT_EXCEL, LEGACY_EXCEL])
    seen: set[str] = set()
    out: list[str] = []
    for p in paths:
        if p not in seen:
            seen.add(p)
            out.append(p)
    return out


def resolve_cache_path() -> str:
    cfg = load_grabber_config()
    if cfg.get("output_json"):
        return os.path.join(SCRIPT_DIR, cfg["output_json"])
    data_json = os.path.join(SCRIPT_DIR, "data", "display_cache.json")
    if os.path.isfile(data_json):
        return data_json
    return CACHE_FILE


def load_sql_query(cfg: dict | None = None, *, path: str | None = None) -> str:
    cfg = build_runtime_config(cfg)
    path = path or cfg["sql_file"]
    if not os.path.isfile(path):
        raise FileNotFoundError(f"找不到 SQL 文件: {path}")
    with open(path, "r", encoding="utf-8") as f:
        lines = [ln for ln in f.readlines() if not ln.strip().startswith("--")]
    query = "\n".join(lines).strip()
    if not query:
        raise ValueError(f"SQL 文件为空: {path}")
    return query


def _is_schema_sql_error(exc: Exception) -> bool:
    msg = str(exc).lower()
    return (
        "invalid object name" in msg
        or "invalid column name" in msg
        or "208," in msg
        or "207," in msg
    )


def _sql_fallback_paths(primary_path: str) -> list[str]:
    folder = os.path.dirname(primary_path) or _resolve_path("sql")
    ordered = [primary_path]
    for name in ("display.sql", "display.minimal.sql"):
        candidate = os.path.join(folder, name)
        if candidate not in ordered and os.path.isfile(candidate):
            ordered.append(candidate)
    return ordered


def last_sql_file() -> str | None:
    return _last_sql_file


def _load_config() -> dict:
    return load_grabber_config()


def parse_mssql_url(url: str) -> dict:
    """从 SQLAlchemy 连接字符串解析服务器/用户名/密码/库名。"""
    url = (url or "").strip()
    if not url:
        return {}
    try:
        from sqlalchemy.engine.url import make_url

        parsed = make_url(url)
        return {
            "db_server": parsed.host or "",
            "db_port": parsed.port or 1433,
            "db_user": parsed.username or "",
            "db_password": parsed.password or "",
            "db_name": (parsed.database or "").lstrip("/"),
        }
    except Exception:
        return {}


def normalize_db_config(cfg: dict) -> dict:
    """合并旧版 database_url 与分项字段，分项优先。"""
    merged = dict(cfg)
    url = (merged.get("database_url") or "").strip()
    if url:
        for key, value in parse_mssql_url(url).items():
            if not str(merged.get(key) or "").strip() and value:
                merged[key] = value
    if not merged.get("db_port"):
        merged["db_port"] = 1433
    return merged


def build_database_url(cfg: dict) -> str | None:
    """优先使用 database_url（与 main_gui 一致）；分项字段仅作备用。"""
    env = os.environ.get("DISPLAY_DB_URL", "").strip()
    if env:
        return env

    url = (cfg.get("database_url") or "").strip()
    if url:
        return url

    merged = normalize_db_config(cfg)
    server = (merged.get("db_server") or "").strip()
    user = (merged.get("db_user") or "").strip()
    password = merged.get("db_password")
    db_name = (merged.get("db_name") or "").strip()
    if server and user and password not in (None, "") and db_name:
        port = int(merged.get("db_port") or 1433)
        enc_user = quote_plus(user)
        enc_pass = quote_plus(str(password))
        return (
            f"mssql+pymssql://{enc_user}:{enc_pass}@{server}:{port}/{db_name}?charset=utf8"
        )
    return None


def resolve_database_url(cfg: dict) -> str:
    """供界面显示：优先 database_url，否则由分项拼出。"""
    url = (cfg.get("database_url") or "").strip()
    if url:
        return url
    built = build_database_url(cfg)
    return built or ""


def format_db_error(exc: Exception) -> str:
    msg = str(exc)
    if "YOUR_PASSWORD" in msg:
        return "请在 grabber_config.json 或界面中填写真实数据库密码（不要用 YOUR_PASSWORD 占位符）。"
    if "18456" in msg or "Login failed" in msg:
        return (
            "数据库登录失败：请从 main_gui 复制完整 database_url 连接串粘贴到此处。"
            "密码中的 @ ^ ! 必须写成 %40 %5E %21 等形式，不能直接写明文。"
        )
    if "40615" in msg or "not allowed to access the server" in msg.lower():
        return "无法连接 Azure SQL：请在防火墙中添加当前公网 IP 后重试。"
    return msg


def test_database_connection(cfg: dict | None = None) -> tuple[bool, str]:
    runtime = build_runtime_config(cfg or {})
    url = build_database_url(runtime)
    if not url:
        return False, "未配置 database_url 连接串。"
    try:
        from sqlalchemy import create_engine, text

        engine = create_engine(url, connect_args={"timeout": 15})
        with engine.connect() as conn:
            conn.execute(text("SELECT 1"))
        return True, "连接成功"
    except Exception as exc:
        return False, format_db_error(exc)


def _database_url(cfg: dict) -> str | None:
    return build_database_url(cfg)


def _fetch_raw_rows(
    cfg: dict,
    query: str | None = None,
    *,
    canonicalize: bool = True,
) -> list[dict]:
    global _last_sql_file
    url = _database_url(cfg)
    if not url:
        raise RuntimeError(
            "未配置 database_url（grabber_config.json 或环境变量 DISPLAY_DB_URL）"
        )

    from sqlalchemy import create_engine, text

    def _row_from_raw(raw: dict) -> dict:
        if canonicalize:
            return _canonicalize_row(raw)
        return dict(raw)

    def _execute(sql_text: str) -> list[dict]:
        engine = create_engine(url, connect_args={"timeout": 30})
        rows: list[dict] = []
        with engine.connect() as conn:
            result = conn.execute(text(sql_text))
            keys = list(result.keys())
            for row in result:
                raw = {keys[i]: row[i] for i in range(len(keys))}
                rows.append(_row_from_raw(raw))
        return rows

    if (query or cfg.get("query") or "").strip():
        sql_text = (query or cfg.get("query") or "").strip()
        try:
            rows = _execute(sql_text)
            _last_sql_file = cfg.get("sql_file")
            return rows
        except Exception as exc:
            raise RuntimeError(format_db_error(exc)) from exc

    primary = cfg["sql_file"]
    last_exc: Exception | None = None
    for path in _sql_fallback_paths(primary):
        try:
            rows = _execute(load_sql_query(cfg, path=path))
            _last_sql_file = path
            return rows
        except Exception as exc:
            if _is_schema_sql_error(exc):
                last_exc = exc
                continue
            raise RuntimeError(format_db_error(exc)) from exc

    hint = "请在 SSMS 运行 sql/discover_schema.sql，把 Products 表的图片列名发给我们。"
    if last_exc is not None:
        raise RuntimeError(f"{format_db_error(last_exc)}\n{hint}") from last_exc
    raise RuntimeError(hint)


def write_rows_to_excel(rows: list[dict], path: str) -> None:
    """通用 SQL 结果导出：保留原始列名（用于周销量等非 Display 数据）。"""
    os.makedirs(os.path.dirname(path) or ".", exist_ok=True)
    if not rows:
        try:
            import pandas as pd

            pd.DataFrame().to_excel(path, index=False)
            return
        except ImportError:
            from openpyxl import Workbook

            Workbook().save(path)
            return

    columns: list[str] = []
    seen: set[str] = set()
    for row in rows:
        for key in row:
            if key not in seen:
                seen.add(key)
                columns.append(key)

    try:
        import pandas as pd

        pd.DataFrame(rows, columns=columns).to_excel(path, index=False)
        return
    except ImportError:
        pass

    from openpyxl import Workbook

    wb = Workbook()
    ws = wb.active
    ws.append(columns)
    for row in rows:
        ws.append([row.get(col) for col in columns])
    wb.save(path)


def grab_sql_to_excel(cfg: dict | None = None) -> tuple[list[dict], str]:
    """通用抓取：cfg 指定 sql_file + output_excel，写入 Excel 并返回原始行。"""
    runtime = build_runtime_config(cfg)
    rows = _fetch_raw_rows(runtime, canonicalize=False)
    excel_path = runtime["output_excel"]
    write_rows_to_excel(rows, excel_path)
    return rows, excel_path


def export_rows_to_excel(rows: list[dict], path: str) -> None:
    os.makedirs(os.path.dirname(path) or ".", exist_ok=True)
    export = []
    for row in rows:
        export.append({
            "WarehouseName": row.get("warehouse_name", ""),
            "Sku": row.get("product_code", ""),
            "ProductName": row.get("product_name", ""),
            "ProductFamily": row.get("product_family", ""),
            "SubProductFamily": row.get("sub_product_family", ""),
            "ImageUrl": row.get("image_url", ""),
            "DisplayQty": row.get("display_qty", 0),
        })
    try:
        import pandas as pd

        pd.DataFrame(export).to_excel(path, index=False)
        return
    except ImportError:
        pass
    from openpyxl import Workbook

    wb = Workbook()
    ws = wb.active
    ws.append(["WarehouseName", "Sku", "ProductName", "ProductFamily", "SubProductFamily", "ImageUrl", "DisplayQty"])
    for r in export:
        ws.append([
            r["WarehouseName"],
            r["Sku"],
            r["ProductName"],
            r["ProductFamily"],
            r["SubProductFamily"],
            r["ImageUrl"],
            r["DisplayQty"],
        ])
    wb.save(path)


def sales_runtime_config(cfg: dict | None = None) -> dict:
    """周销量抓取路径（与 Display 共用 database_url）。"""
    base = build_runtime_config(cfg)
    sql_file = base.get("sales_sql_file") or os.path.join(
        base.get("sql_folder") or "sql", "weekly_sales.sql"
    )
    if not os.path.isabs(sql_file):
        sql_file = os.path.join(SCRIPT_DIR, sql_file)
    output_excel = base.get("sales_output_excel") or os.path.join(
        base.get("output_folder") or "data", "weekly_sales.xlsx"
    )
    if not os.path.isabs(output_excel):
        output_excel = os.path.join(SCRIPT_DIR, output_excel)
    return {**base, "sql_file": sql_file, "output_excel": output_excel}


def grab_weekly_sales(cfg: dict | None = None) -> tuple[list[dict], str]:
    """周销量抓取：SQL → data/weekly_sales.xlsx。"""
    runtime = sales_runtime_config(cfg)
    rows, excel_path = grab_sql_to_excel(runtime)
    try:
        from sales_lookup import reload_weekly_sales

        reload_weekly_sales(excel_path)
    except Exception:
        pass
    return rows, excel_path


def run_grab_pipeline(
    cfg: dict | None = None,
    *,
    display: bool = True,
    sales: bool = False,
    sync_roi: bool = False,
    log=print,
) -> dict:
    """统一抓取：Display + 周销量，可选同步 ROI 到模板。"""
    results: dict = {}
    if display:
        log("抓取 Display 数据...")
        items, excel_path = grab_and_save(cfg)
        results["display"] = {"items": items, "excel": excel_path, "count": len(items)}
        log(f"Display 完成: {len(items)} 款 → {excel_path}")
    if sales:
        log("抓取周销量...")
        rows, excel_path = grab_weekly_sales(cfg)
        results["sales"] = {"rows": rows, "excel": excel_path, "count": len(rows)}
        log(f"周销量完成: {len(rows)} 行 → {excel_path}")
    if sync_roi:
        log("同步 ROI 到模板与布局...")
        try:
            import importlib.util

            roi_script = os.path.join(SCRIPT_DIR, "scripts", "update_roi.py")
            spec = importlib.util.spec_from_file_location("update_roi", roi_script)
            if spec is None or spec.loader is None:
                raise ImportError(f"无法加载 {roi_script}")
            mod = importlib.util.module_from_spec(spec)
            spec.loader.exec_module(mod)
            code = int(mod.main())
            results["roi_sync"] = {"ok": code == 0}
            if code == 0:
                log("ROI 同步完成")
            else:
                log("ROI 同步未完全成功，请查看日志")
        except Exception as exc:
            results["roi_sync"] = {"ok": False, "error": str(exc)}
            log(f"ROI 同步失败: {exc}")
    return results


def grab_and_save(cfg: dict | None = None) -> tuple[list["DisplayItem"], str]:
    """Main 抓取：SQL → data/display.xlsx + JSON 缓存。"""
    global _display_cache, _last_load_error, _last_load_source
    runtime = build_runtime_config(cfg)
    rows = _fetch_raw_rows(runtime)
    excel_path = runtime["output_excel"]
    export_rows_to_excel(rows, excel_path)
    items_all = _rows_to_items(rows, "warehouse", apply_blacklist=False)
    global _display_items_all
    _display_items_all = items_all
    items = filter_blacklisted(items_all)
    save_cache(items, runtime["output_json"])
    _display_cache = items
    _last_load_error = None
    _last_load_source = os.path.basename(excel_path)
    return items, excel_path


def _fetch_from_database(cfg: dict) -> list[DisplayItem]:
    rows = _fetch_raw_rows(cfg)
    fmt = _detect_format(rows[0]) if rows else "warehouse"
    return _rows_to_items(rows, fmt)


def _load_cache_file(path: str | None = None) -> list[DisplayItem]:
    path = path or CACHE_FILE
    if not os.path.isfile(path):
        return []
    with open(path, "r", encoding="utf-8") as f:
        payload = json.load(f)
    raw = payload.get("items", payload if isinstance(payload, list) else [])
    return _rows_to_items(raw)


def save_cache(items: list[DisplayItem], path: str | None = None) -> None:
    path = path or CACHE_FILE
    payload = {
        "updated_at": datetime.now(timezone.utc).isoformat(),
        "items": [
            {
                "product_code": it.product_code,
                "product_name": it.product_name,
                "product_family": it.product_family,
                "sub_product_family": it.sub_product_family,
                "image_url": it.image_url,
                "stock_details": it.stock_details,
                "displays": [
                    {
                        "shop_id": s.shop_id,
                        "shop_label": s.shop_label,
                        "location": s.location,
                        "qty": s.qty,
                    }
                    for s in it.displays
                ],
            }
            for it in items
        ],
    }
    with open(path, "w", encoding="utf-8") as f:
        json.dump(payload, f, ensure_ascii=False, indent=2)


def refresh_from_database() -> list[DisplayItem]:
    global _display_cache, _last_load_error, _last_load_source
    cfg = _load_config()
    items = _fetch_from_database(cfg)
    save_cache(items)
    _display_cache = items
    _last_load_error = None
    _last_load_source = "database"
    return items


def _display_cache_is_fresh(cache_path: str, excel_path: str | None) -> bool:
    if not excel_path or not os.path.isfile(excel_path) or not os.path.isfile(cache_path):
        return False
    try:
        return os.path.getmtime(cache_path) >= os.path.getmtime(excel_path)
    except OSError:
        return False


def load_display_items(*, prefer_db: bool = False) -> list[DisplayItem]:
    """可视化程序读取：JSON 缓存（若较新）→ Excel → 数据库抓取。"""
    global _display_cache, _last_load_error, _last_load_source
    if _display_cache is not None and not prefer_db:
        return _display_cache

    if prefer_db:
        try:
            items, _ = grab_and_save()
            return items
        except Exception as exc:
            _last_load_error = str(exc)
            print(f"Display 抓取失败，尝试读本地 Excel: {exc}")

    cache_path = resolve_cache_path()
    excel_path = next((p for p in resolve_display_excel_paths() if os.path.isfile(p)), None)
    if _display_cache_is_fresh(cache_path, excel_path):
        items = _load_cache_file(cache_path)
        if items:
            _display_cache = items
            _last_load_error = None
            _last_load_source = os.path.basename(cache_path)
            return items

    items = load_from_excel()
    if items:
        _display_cache = items
        return items

    items = _load_cache_file(cache_path)
    _display_cache = items
    if items:
        _last_load_source = os.path.basename(resolve_cache_path())
        return items

    if _last_load_error is None:
        _last_load_error = "请先运行 grab_display.bat 抓取数据"
    return items


def reload_display_items(*, prefer_db: bool = False) -> list[DisplayItem]:
    global _display_cache
    _display_cache = None
    return load_display_items(prefer_db=prefer_db)


def last_load_error() -> str | None:
    return _last_load_error


def last_load_source() -> str | None:
    return _last_load_source


def last_family_column() -> str | None:
    return _last_family_column


def filter_gallery_items(
    items: list[DisplayItem],
    shop_id: str,
    query: str,
    templates: list[dict],
    *,
    survey_filter: str = "all",
    blacklist_mode: str = "exclude",
) -> list[DisplayItem]:
    """画廊筛选：门店 / 搜索 / 已测绘 / 黑名单模式。"""
    blocked = load_blacklist()
    q = _normalize_key(query)
    out: list[DisplayItem] = []
    for it in items:
        bl = is_blacklisted(it, blocked)
        if blacklist_mode == "exclude" and bl:
            continue
        if blacklist_mode == "only" and not bl:
            continue
        if shop_id != "all" and it.display_qty_for_shop(shop_id) <= 0:
            continue
        if q:
            blob = " ".join([
                it.product_code,
                it.product_name,
                it.product_family,
                it.sub_product_family,
            ]).lower()
            if q not in blob:
                continue
        modeled = match_template_index(it, templates) >= 0
        if survey_filter == "modeled" and not modeled:
            continue
        if survey_filter == "unmodeled" and modeled:
            continue
        out.append(it)
    out.sort(key=lambda x: (
        x.product_family.lower(),
        x.sub_product_family.lower(),
        x.product_name.lower(),
    ))
    return out


def display_items_including_blacklist() -> list[DisplayItem]:
    """含黑名单的完整 Display 列表（画廊「全部/仅黑名单」用）。"""
    if _display_items_all is not None:
        return _display_items_all
    return _display_cache or []


def filter_items(items: list[DisplayItem], shop_id: str, query: str = "") -> list[DisplayItem]:
    q = _normalize_key(query)
    out: list[DisplayItem] = []
    for it in items:
        if shop_id != "all" and it.display_qty_for_shop(shop_id) <= 0:
            continue
        if q:
            blob = " ".join([
                it.product_code,
                it.product_name,
                it.product_family,
                it.sub_product_family,
            ]).lower()
            if q not in blob:
                continue
        out.append(it)
    out.sort(key=lambda x: (
        x.product_family.lower(),
        x.sub_product_family.lower(),
        x.product_name.lower(),
    ))
    return out


def group_by_family(items: list[DisplayItem]) -> list[tuple[str, list[DisplayItem]]]:
    groups: dict[str, list[DisplayItem]] = {}
    for it in items:
        groups.setdefault(it.product_family or "未分类", []).append(it)
    return sorted(groups.items(), key=lambda x: (-len(x[1]), x[0].lower()))


def group_by_family_hierarchy(
    items: list[DisplayItem],
) -> list[tuple[str, list[tuple[str, list[DisplayItem]]]]]:
    """Product Family → Sub Product Family → items。"""
    top: dict[str, dict[str, list[DisplayItem]]] = {}
    for it in items:
        fam = it.product_family or "未分类"
        sub = it.sub_product_family or it.product_name or it.product_code or "未分类"
        top.setdefault(fam, {}).setdefault(sub, []).append(it)
    out: list[tuple[str, list[tuple[str, list[DisplayItem]]]]] = []
    for fam in sorted(top.keys(), key=str.lower):
        subs = sorted(top[fam].items(), key=lambda x: x[0].lower())
        out.append((fam, subs))
    return out


def shop_stats(items: list[DisplayItem], templates: list[dict]) -> dict[str, dict[str, int]]:
    stats: dict[str, dict[str, int]] = {}
    for shop in SHOPS:
        sid = shop["id"]
        filtered = filter_items(items, sid)
        modeled = sum(1 for it in filtered if match_template_index(it, templates) >= 0)
        families = len({
            it.product_family
            for it in filtered
            if it.product_family and it.product_family != "未分类"
        })
        stats[sid] = {"total": len(filtered), "modeled": modeled, "families": families}
    return stats


def shops_for_display_tabs(
    items: list[DisplayItem], templates: list[dict]
) -> list[tuple[dict[str, Any], dict[str, int]]]:
    """门店 Tab：全部固定第一，其余按 Product Family 数量从高到低。"""
    stats = shop_stats(items, templates)
    rows: list[tuple[dict[str, Any], dict[str, int]]] = []
    for shop in SHOPS:
        sid = shop["id"]
        if sid == "other":
            continue
        st = stats.get(sid, {"total": 0, "modeled": 0, "families": 0})
        if sid != "all" and st["total"] == 0:
            continue
        rows.append((shop, st))
    all_row = next((r for r in rows if r[0]["id"] == "all"), None)
    rest = [r for r in rows if r[0]["id"] != "all"]
    rest.sort(key=lambda r: (-r[1].get("families", 0), -r[1]["total"], r[0]["label"].lower()))
    out: list[tuple[dict[str, Any], dict[str, int]]] = []
    if all_row:
        out.append(all_row)
    out.extend(rest)
    return out


def match_template_index(item: DisplayItem, templates: list[dict]) -> int:
    """按 SKU 或产品名精确匹配模板，不按 Family 模糊匹配（避免测绘一款整族都变绿）。"""
    code = _normalize_key(item.product_code)
    name = _normalize_key(item.product_name)
    for i, tpl in enumerate(templates):
        tid = _normalize_key(tpl.get("id", ""))
        if code and code == tid:
            return i
        if name and name == tid:
            return i
    return -1


def find_template_index_by_id(templates: list[dict], tpl_id: str) -> int:
    key = _normalize_key(tpl_id)
    for i, tpl in enumerate(templates):
        if _normalize_key(tpl.get("id", "")) == key:
            return i
    return -1


def template_matches_display_item(tpl: dict, item: DisplayItem) -> bool:
    """模板 id 是否与 Display 产品的 SKU 或产品名一致。"""
    tid = _normalize_key(tpl.get("id", ""))
    if not tid:
        return False
    code = _normalize_key(item.product_code)
    name = _normalize_key(item.product_name)
    return (code and code == tid) or (name and name == tid)


def find_display_item_for_template(
    tpl: dict, items: list[DisplayItem]
) -> DisplayItem | None:
    for item in items:
        if template_matches_display_item(tpl, item):
            return item
    return None


def prune_orphan_templates(
    templates: list[dict], items: list[DisplayItem]
) -> tuple[list[dict], list[str]]:
    """只保留能在 Display 库中匹配到的模板，删除手工测绘的游离项。"""
    if not items:
        return templates[:], []
    kept: list[dict] = []
    removed: list[str] = []
    for tpl in templates:
        if find_display_item_for_template(tpl, items):
            kept.append(tpl)
        else:
            removed.append(str(tpl.get("id", "")))
    return kept, removed
