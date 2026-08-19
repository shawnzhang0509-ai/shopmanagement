"""从 roi.xlsx 或周销量 Excel 按 product_family 查询 ROI。"""
from __future__ import annotations

import os

ROI_FILE = "roi.xlsx"
_roi_cache: dict[str, float] | None = None


def _normalize_key(value) -> str:
    return str(value).strip().lower()


def _load_with_openpyxl() -> dict[str, float]:
    from openpyxl import load_workbook

    result: dict[str, float] = {}
    wb = load_workbook(ROI_FILE, read_only=True, data_only=True)
    ws = wb.active
    rows = list(ws.iter_rows(values_only=True))
    if not rows:
        return result
    headers = [str(h).strip().lower() if h is not None else "" for h in rows[0]]
    family_col = None
    for candidate in ("product_family", "product family", "family", "id"):
        if candidate in headers:
            family_col = headers.index(candidate)
            break
    if family_col is None or "roi" not in headers:
        return result
    roi_col = headers.index("roi")
    for row in rows[1:]:
        if not row or family_col >= len(row) or roi_col >= len(row):
            continue
        key = _normalize_key(row[family_col])
        roi = row[roi_col]
        if key and roi is not None:
            try:
                result[key] = float(roi)
            except (TypeError, ValueError):
                pass
    wb.close()
    return result


def _load_with_pandas() -> dict[str, float]:
    import pandas as pd

    result: dict[str, float] = {}
    df = pd.read_excel(ROI_FILE)
    family_col = None
    for col in ("product_family", "product family", "family", "id"):
        if col in df.columns:
            family_col = col
            break
    if family_col is None or "roi" not in df.columns:
        return result
    for _, row in df.iterrows():
        key = _normalize_key(row[family_col])
        roi = row["roi"]
        if key and pd.notna(roi):
            result[key] = float(roi)
    return result


def load_roi_map() -> dict[str, float]:
    global _roi_cache
    if _roi_cache is not None:
        return _roi_cache

    _roi_cache = {}
    if not os.path.isfile(ROI_FILE):
        print(f"未找到 {ROI_FILE}")
        return _roi_cache

    try:
        try:
            import pandas as pd  # noqa: F401
            _roi_cache = _load_with_pandas()
        except ImportError:
            _roi_cache = _load_with_openpyxl()
        if not _roi_cache:
            _roi_cache = _load_with_openpyxl()
    except Exception as exc:
        print(f"读取 ROI 表失败: {exc}")

    return _roi_cache


def lookup_roi(product_family: str, shop_id: str | None = None) -> float:
    if not product_family:
        return 0.0

    try:
        from sales_lookup import resolve_product_family

        family = resolve_product_family(product_family, shop_id=shop_id)
    except Exception:
        family = product_family

    roi_map = load_roi_map()
    static = roi_map.get(_normalize_key(family), 0.0)
    if static <= 0:
        static = roi_map.get(_normalize_key(product_family), 0.0)
    if static > 0:
        return static

    try:
        from sales_lookup import lookup_sales_roi, lookup_sales_roi_by_sku, sales_data_available

        if not sales_data_available():
            return 0.0
        roi = lookup_sales_roi(family, shop_id)
        if roi > 0:
            return roi
        if family != product_family:
            roi = lookup_sales_roi(product_family, shop_id)
            if roi > 0:
                return roi
        roi = lookup_sales_roi_by_sku(product_family, shop_id)
        if roi > 0:
            return roi
        if family != product_family:
            return lookup_sales_roi_by_sku(family, shop_id)
    except Exception:
        pass

    return 0.0


def resolve_furniture_roi(
    name: str,
    product_family: str = "",
    shop_id: str | None = None,
) -> tuple[str, float]:
    """解析 SKU/系列并返回 (product_family, roi)。"""
    raw = (product_family or name or "").strip()
    try:
        from sales_lookup import resolve_product_family

        family = resolve_product_family(raw, shop_id=shop_id) or raw
    except Exception:
        family = raw
    roi = lookup_roi(family, shop_id)
    if roi <= 0 and raw and _normalize_key(raw) != _normalize_key(family):
        roi = lookup_roi(raw, shop_id)
    return family, roi


def reload_roi_map() -> dict[str, float]:
    global _roi_cache
    _roi_cache = None
    return load_roi_map()
