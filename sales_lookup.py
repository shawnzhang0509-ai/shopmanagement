"""周销量数据层：grab_sales.bat → data/weekly_sales.xlsx → ROI / 坪效分析。

与 display_lookup（Display 库存）分离：
  grab_display.bat  → sql/display.sql      → data/display.xlsx
  grab_sales.bat    → sql/weekly_sales.sql  → data/weekly_sales.xlsx
"""
from __future__ import annotations

import os
import re
from dataclasses import dataclass
from typing import Any

from display_lookup import SCRIPT_DIR, shop_id_for_location

DEFAULT_EXCEL = os.path.join(SCRIPT_DIR, "data", "weekly_sales.xlsx")
SALES_SQL = os.path.join(SCRIPT_DIR, "sql", "weekly_sales.sql")

_sales_cache: list["WeeklySalesRow"] | None = None
_family_totals_cache: dict[tuple[str | None, str], dict[str, float]] = {}
_sku_totals_cache: dict[tuple[str | None, str], dict[str, float]] = {}


def _normalize_key(value: Any) -> str:
    return re.sub(r"\s+", " ", str(value or "").strip().lower())


def _to_float(value: Any) -> float:
    if value is None:
        return 0.0
    try:
        return float(value)
    except (TypeError, ValueError):
        return 0.0


def _to_int(value: Any) -> int:
    return int(_to_float(value))


@dataclass
class WeeklySalesRow:
    branch_name: str
    shop_id: str
    product_family: str
    channel: str
    sku: str
    product_name: str
    year_week_period: str
    total_qty: float
    total_amount: float
    order_count: int
    avg_unit_price: float


def _header_map(headers: list[str]) -> dict[str, int]:
    normalized = {_normalize_key(h): i for i, h in enumerate(headers)}
    aliases = {
        "branch_name": ("branchname", "branch", "store", "warehouse"),
        "product_family": ("productfamily", "family", "product family"),
        "channel": ("channel",),
        "sku": ("sku", "product_code", "productcode"),
        "product_name": ("productname", "name", "product name"),
        "year_week_period": ("yearweekperiod", "yearweek", "week", "year week"),
        "total_qty": ("totalqty", "qty", "quantity", "销量"),
        "total_amount": ("totalamount", "amount", "sales", "销售金额"),
        "order_count": ("ordercount", "orders", "订单笔数"),
        "avg_unit_price": ("avgunitprice", "unitprice", "平均单价"),
    }
    out: dict[str, int] = {}
    for field, keys in aliases.items():
        for key in keys:
            if key in normalized:
                out[field] = normalized[key]
                break
    return out


def _row_value(row: tuple, col_map: dict[str, int], field: str, default: Any = "") -> Any:
    idx = col_map.get(field)
    if idx is None or idx >= len(row):
        return default
    return row[idx]


def _rows_from_openpyxl(path: str) -> list[WeeklySalesRow]:
    from openpyxl import load_workbook

    wb = load_workbook(path, read_only=True, data_only=True)
    ws = wb.active
    rows = list(ws.iter_rows(values_only=True))
    wb.close()
    if not rows:
        return []
    col_map = _header_map([str(h or "") for h in rows[0]])
    if "product_family" not in col_map:
        return []
    result: list[WeeklySalesRow] = []
    for row in rows[1:]:
        if not row:
            continue
        family = str(_row_value(row, col_map, "product_family", "") or "").strip()
        if not family:
            continue
        branch = str(_row_value(row, col_map, "branch_name", "") or "").strip()
        result.append(
            WeeklySalesRow(
                branch_name=branch,
                shop_id=shop_id_for_location(branch),
                product_family=family,
                channel=str(_row_value(row, col_map, "channel", "") or ""),
                sku=str(_row_value(row, col_map, "sku", "") or ""),
                product_name=str(_row_value(row, col_map, "product_name", "") or ""),
                year_week_period=str(_row_value(row, col_map, "year_week_period", "") or ""),
                total_qty=_to_float(_row_value(row, col_map, "total_qty", 0)),
                total_amount=_to_float(_row_value(row, col_map, "total_amount", 0)),
                order_count=_to_int(_row_value(row, col_map, "order_count", 0)),
                avg_unit_price=_to_float(_row_value(row, col_map, "avg_unit_price", 0)),
            )
        )
    return result


def _rows_from_pandas(path: str) -> list[WeeklySalesRow]:
    import pandas as pd

    df = pd.read_excel(path)
    if df.empty:
        return []
    col_map = _header_map([str(c) for c in df.columns])
    if "product_family" not in col_map:
        return []
    result: list[WeeklySalesRow] = []
    for row in df.itertuples(index=False, name=None):
        family = str(_row_value(row, col_map, "product_family", "") or "").strip()
        if not family:
            continue
        branch = str(_row_value(row, col_map, "branch_name", "") or "").strip()
        result.append(
            WeeklySalesRow(
                branch_name=branch,
                shop_id=shop_id_for_location(branch),
                product_family=family,
                channel=str(_row_value(row, col_map, "channel", "") or ""),
                sku=str(_row_value(row, col_map, "sku", "") or ""),
                product_name=str(_row_value(row, col_map, "product_name", "") or ""),
                year_week_period=str(_row_value(row, col_map, "year_week_period", "") or ""),
                total_qty=_to_float(_row_value(row, col_map, "total_qty", 0)),
                total_amount=_to_float(_row_value(row, col_map, "total_amount", 0)),
                order_count=_to_int(_row_value(row, col_map, "order_count", 0)),
                avg_unit_price=_to_float(_row_value(row, col_map, "avg_unit_price", 0)),
            )
        )
    return result


def load_weekly_sales(path: str | None = None) -> list[WeeklySalesRow]:
    global _sales_cache
    if _sales_cache is not None and path is None:
        return _sales_cache

    path = path or DEFAULT_EXCEL
    if not os.path.isfile(path):
        _sales_cache = []
        return _sales_cache

    try:
        try:
            import pandas as pd  # noqa: F401

            rows = _rows_from_pandas(path)
        except ImportError:
            rows = _rows_from_openpyxl(path)
        if not rows:
            rows = _rows_from_openpyxl(path)
    except Exception as exc:
        print(f"读取周销量表失败: {exc}")
        rows = []

    if path is None or path == DEFAULT_EXCEL:
        _sales_cache = rows
    return rows


_sku_family_map: dict[str, str] | None = None


def sku_family_map(*, path: str | None = None) -> dict[str, str]:
    """Sku → ProductFamily，来自 weekly_sales.xlsx，辅以 display.xlsx。"""
    global _sku_family_map
    if _sku_family_map is not None and path is None:
        return _sku_family_map

    mapping: dict[str, str] = {}
    for row in load_weekly_sales(path):
        sku = _normalize_key(row.sku)
        fam = str(row.product_family or "").strip()
        if sku and fam:
            mapping[sku] = fam

    try:
        from display_lookup import load_display_items

        for item in load_display_items():
            sku = _normalize_key(item.product_code)
            fam = str(item.product_family or "").strip()
            if sku and fam:
                mapping.setdefault(sku, fam)
    except Exception:
        pass

    if path is None:
        _sku_family_map = mapping
    return mapping


def resolve_product_family(key: str, *, shop_id: str | None = None) -> str:
    """把 SKU / 系列名统一解析为 ProductFamily。"""
    text = str(key or "").strip()
    if not text:
        return ""
    norm = _normalize_key(text)
    mapped = sku_family_map().get(norm)
    if mapped:
        return mapped
    if norm in aggregate_by_family(shop_id):
        return text
    return text


def reload_weekly_sales(path: str | None = None) -> list[WeeklySalesRow]:
    global _sales_cache, _family_totals_cache, _sku_family_map, _sku_totals_cache
    _sales_cache = None
    _family_totals_cache = {}
    _sku_family_map = None
    _sku_totals_cache = {}
    return load_weekly_sales(path)


def aggregate_by_family(
    shop_id: str | None = None,
    *,
    path: str | None = None,
) -> dict[str, dict[str, float]]:
    """按 ProductFamily 汇总销量/金额（可选按门店 shop_id 过滤）。"""
    cache_key = (shop_id, path or DEFAULT_EXCEL)
    if cache_key in _family_totals_cache:
        return _family_totals_cache[cache_key]

    totals: dict[str, dict[str, float]] = {}
    for row in load_weekly_sales(path):
        if shop_id and shop_id not in ("all", "") and row.shop_id != shop_id:
            continue
        key = _normalize_key(row.product_family)
        bucket = totals.setdefault(
            key,
            {"total_qty": 0.0, "total_amount": 0.0, "order_count": 0.0},
        )
        bucket["total_qty"] += row.total_qty
        bucket["total_amount"] += row.total_amount
        bucket["order_count"] += row.order_count

    _family_totals_cache[cache_key] = totals
    return totals


def aggregate_by_sku(
    shop_id: str | None = None,
    *,
    path: str | None = None,
) -> dict[str, dict[str, float]]:
    """按 Sku 汇总销量/金额（可选按门店 shop_id 过滤）。"""
    cache_key = (shop_id, path or DEFAULT_EXCEL)
    if cache_key in _sku_totals_cache:
        return _sku_totals_cache[cache_key]

    totals: dict[str, dict[str, float]] = {}
    for row in load_weekly_sales(path):
        if shop_id and shop_id not in ("all", "") and row.shop_id != shop_id:
            continue
        sku = _normalize_key(row.sku)
        if not sku:
            continue
        bucket = totals.setdefault(
            sku,
            {"total_qty": 0.0, "total_amount": 0.0, "order_count": 0.0, "product_family": row.product_family},
        )
        bucket["total_qty"] += row.total_qty
        bucket["total_amount"] += row.total_amount
        bucket["order_count"] += row.order_count
        if row.product_family and not bucket.get("product_family"):
            bucket["product_family"] = row.product_family

    _sku_totals_cache[cache_key] = totals
    return totals


def lookup_sales_roi_by_sku(sku: str, shop_id: str | None = None) -> float:
    if not sku:
        return 0.0
    totals = aggregate_by_sku(shop_id)
    if not totals:
        return 0.0
    key = _normalize_key(sku)
    mine = totals.get(key)
    if not mine or mine["total_amount"] <= 0:
        return 0.0
    max_amount = max(v["total_amount"] for v in totals.values() if v["total_amount"] > 0)
    if max_amount <= 0:
        return 0.0
    return min(10.0, mine["total_amount"] / max_amount * 10.0)


def lookup_sales_roi(product_family: str, shop_id: str | None = None) -> float:
    """由周销量金额推算 0–10 ROI（同门店内按金额线性归一化）。

    若你有固定 ROI 公式，可在此函数内替换算法，或继续用 roi.xlsx 覆盖。
    """
    if not product_family:
        return 0.0
    totals = aggregate_by_family(shop_id)
    if not totals:
        return 0.0
    key = _normalize_key(product_family)
    mine = totals.get(key)
    if not mine or mine["total_amount"] <= 0:
        return 0.0
    max_amount = max(v["total_amount"] for v in totals.values() if v["total_amount"] > 0)
    if max_amount <= 0:
        return 0.0
    return min(10.0, mine["total_amount"] / max_amount * 10.0)


def sales_data_available(path: str | None = None) -> bool:
    path = path or DEFAULT_EXCEL
    return os.path.isfile(path) and os.path.getsize(path) > 0
